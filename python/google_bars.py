import time
import googlemaps
from decouple import config
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

gmaps = googlemaps.Client(key=config('API_KEY'))

field_names = {
    'Name of Bar': 'name',
    'Vicinity': 'vicinity',
    'Full Address': 'formatted_address',
    'Phone Number': 'formatted_phone_number',
    'Website': 'website',
    'Price Level': 'price_level',
    'Rating': 'rating',
    'Total User Ratings': 'user_ratings_total',
    'Google Maps Page': 'url',
    'Business Status': 'business_status',
    'Hours of Operation': 'opening_hours'
}

excel_cols = list(field_names.keys())[0:len(field_names) - 1]
excel_cols.extend([
    'Monday Hours', 'Tuesday Hours', 'Wednesday Hours', 'Thursday Hours',
    'Friday Hours', 'Saturday Hours', 'Sunday Hours'
])

cities = [
    'Asheville, NC', 'Atlanta, GA', 'Austin, TX', 'Chicago, IL', 'Dallas, TX',
    'Las Vegas, NV', 'Los Angeles, CA', 'Miami, FL', 'Nashville, TN',
    'New Orleans, LA', 'New York, NY', 'San Francisco, CA'
]

num_cols = len(excel_cols)
num_rows = 61


def geocode(city):
  time.sleep(2)
  print('Retrieving geocode for {}...'.format(city))
  return gmaps.geocode(city)[0]['geometry']['location']


def places_in_city(city, city_geocode, radius=24140):
  def places_details(results):
    place_details_list = []
    for place in results:
      time.sleep(2)
      print('Retrieving details for {}...'.format(place['name']))
      place_details = gmaps.place(place_id=place['place_id'],
                                  fields=list(field_names.values()))
      if 'result' in place_details:
        place_details_list.append(place_details['result'])
    return place_details_list

  places = []
  print('Retrieving page 1 for {}...'.format(city))
  page = gmaps.places(query=f'best local bars in {city}',
                      location=city_geocode,
                      radius=radius,
                      type='bar')
  places.extend(places_details(page['results']))
  i = 0
  while 'next_page_token' in page and i < 4:
    time.sleep(2)
    print('Retrieving page {} for {}...'.format(i + 2, city))
    page = gmaps.places(query=f'best local bars in {city}',
                        location=city_geocode,
                        radius=radius,
                        type='bar',
                        page_token=page['next_page_token'])
    places.extend(places_details(page['results']))
    i += 1
  return places


def desired_fields(place):
  only_hours = lambda place_hours: place_hours[place_hours.find(':') + 2:]

  desired_place = {}

  # handle all other fields besides hours
  for key, val in field_names.items():
    if val != 'photo' and val != 'opening_hours':
      desired_place[key] = place.get(val, '')

  # handle hours
  place_hours = place.get('opening_hours', {}).get('weekday_text', [])
  if len(place_hours) == 7:
    desired_place['Monday Hours'] = only_hours(place_hours[0])
    desired_place['Tuesday Hours'] = only_hours(place_hours[1])
    desired_place['Wednesday Hours'] = only_hours(place_hours[2])
    desired_place['Thursday Hours'] = only_hours(place_hours[3])
    desired_place['Friday Hours'] = only_hours(place_hours[4])
    desired_place['Saturday Hours'] = only_hours(place_hours[5])
    desired_place['Sunday Hours'] = only_hours(place_hours[6])

  return desired_place


def get_geocodes_and_write_excel():
  city_geocodes = list(map(geocode, cities))

  for i, city in enumerate(cities):
    wb = Workbook()

    dest_filename = f'Bars in {city}.xlsx'

    ws = wb.active
    ws.title = f'Bars in {city}'

    for col in range(num_cols):
      cell = ws[f'{get_column_letter(col + 1)}1']
      cell.font = Font(bold=True)
      cell.value = excel_cols[col]

    bars = list(map(desired_fields, places_in_city(city, city_geocodes[i])))

    for j, bar in enumerate(bars):
      col = 1
      for val in bar.values():
        cell = ws[f'{get_column_letter(col)}{j + 2}']
        if str(val).startswith('http'):
          cell.font = Font(color='0000FF')
          cell.hyperlink = val
        cell.value = val
        col += 1

    for column_cells in ws.columns:
      length = max(len(str(cell.value)) for cell in column_cells)
      ws.column_dimensions[column_cells[0].column_letter].width = length

    wb.save(filename=dest_filename)


get_geocodes_and_write_excel()